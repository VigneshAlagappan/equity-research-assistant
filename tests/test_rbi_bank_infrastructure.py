"""sources/rbi_bank_infrastructure.py — parsing RBI's monthly ATM/NEFT/RTGS
bank-level bulletins. Tests build small synthetic workbooks mirroring both
real shapes (ATM's deep nested header + section-header rows; NEFT's simpler
2-row header with no separate index row) rather than depending on the real
downloaded files."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from sources.rbi_bank_infrastructure import parse_bank_infrastructure_file


def _make_atm_like(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "June 2026"
    ws.append([None] * 8)
    ws.append([None, "ATM, Acceptance Infrastructure and Card Statistics for the Month of June 2026"])
    ws.append([None, "Sr. No.", "Bank Name", "Infrastructure", None, "Card Payments"])
    ws.append([None, None, None, "Number - Outstanding (as on month end)", None, "Volume"])
    ws.append([None, None, None, "ATMs & CRMs", "PoS", None])
    ws.append([None, None, None, 1, 2, 3])
    ws.append([None, None, "Scheduled Commercial Banks", None, None, None])
    ws.append([None, None, "Public Sector Banks", None, None, None])
    ws.append([None, 1, "BANK OF BARODA", 9125, 2504, 3402865])
    ws.append([None, 2, "BANK OF INDIA", 5225, 2460, 1481750])
    ws.merge_cells("D4:E4")
    wb.save(path)


def _make_neft_like(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NEFT"
    ws.append([None] * 5)
    ws.append([None, "NATIONAL ELECTRONIC FUNDS TRANSFER (NEFT) - JUNE 2026"])
    ws.append([None, "Sl. No.", "BANK NAME", "RECEIVED INWARD CREDITS", None])
    ws.append([None, None, None, "NO. OF INWARD TRANSACTIONS", "AMOUNT \n(Rs. Crore)"])
    ws.append([None, 1, "ABHYUDAYA CO-OP BANK LTD", 403786, 903.006473192])
    ws.append([None, 2, "AHMEDABAD MERCANTILE COOP BANK", 83095, 468.855762443])
    ws.merge_cells("D3:E3")
    wb.save(path)


@pytest.fixture
def atm_path(tmp_path: Path) -> Path:
    path = tmp_path / "ATMJUNE2026.XLSX"
    _make_atm_like(path)
    return path


@pytest.fixture
def neft_path(tmp_path: Path) -> Path:
    path = tmp_path / "NEFTRTGS062026.XLSX"
    _make_neft_like(path)
    return path


def test_extracts_period_from_title(atm_path: Path) -> None:
    obs = parse_bank_infrastructure_file(atm_path)
    assert obs and all(o.period == "2026-06" for o in obs)
    assert all(o.period_type == "monthly" for o in obs)


def test_section_header_rows_are_not_treated_as_banks(atm_path: Path) -> None:
    obs = parse_bank_infrastructure_file(atm_path)
    banks = {o.bank_name for o in obs}
    assert "Scheduled Commercial Banks" not in banks
    assert "Public Sector Banks" not in banks
    assert banks == {"BANK OF BARODA", "BANK OF INDIA"}


def test_index_row_numbers_are_excluded_from_metric_labels(atm_path: Path) -> None:
    obs = parse_bank_infrastructure_file(atm_path)
    metrics = {o.metric for o in obs}
    assert not any(m.endswith("_1") or m.endswith("_2") or m.endswith("_3") for m in metrics)


def test_merged_header_cells_are_forward_filled(atm_path: Path) -> None:
    """D4:E4 is merged ("Number - Outstanding...") — both the ATMs & CRMs
    and PoS columns must inherit that text, not just the top-left cell."""
    obs = parse_bank_infrastructure_file(atm_path)
    metrics = {o.metric for o in obs}
    assert any("outstanding" in m for m in metrics)
    pos_metrics = [m for m in metrics if "pos" in m]
    assert any("outstanding" in m for m in pos_metrics)


def test_neft_file_with_no_index_row_still_parses(neft_path: Path) -> None:
    obs = parse_bank_infrastructure_file(neft_path)
    banks = {o.bank_name for o in obs}
    assert banks == {"ABHYUDAYA CO-OP BANK LTD", "AHMEDABAD MERCANTILE COOP BANK"}
    assert len(obs) == 4  # 2 banks x 2 metrics


def test_neft_units_inferred_from_header(neft_path: Path) -> None:
    obs = parse_bank_infrastructure_file(neft_path)
    amount = next(o for o in obs if "amount" in o.metric)
    count = next(o for o in obs if "transactions" in o.metric)
    assert amount.unit == "INR_CRORE"
    assert count.unit == "NUMBER"


def test_raises_for_a_file_with_no_bank_name_column(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.active.append(["just", "some", "random", "cells"])
    path = tmp_path / "not_a_bulletin.xlsx"
    wb.save(path)

    with pytest.raises(ValueError):
        parse_bank_infrastructure_file(path)
