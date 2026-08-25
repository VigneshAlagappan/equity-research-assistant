"""End-to-end: detect -> parse -> validate -> store -> reconcile.

Exercises the trivial pass-through reconciliation the README describes for
Phase 2 (README: Data Layers -> Raw -> Normalized -> Derived, worked example):
a single source means canonical_financials just points at that one observation,
with reconciliation_reason "only source available".
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest

from companies.lifecycle import archive_company
from companies.registry import seed_companies
from ingestion.detector import (
    PathConventionError,
    detect_from_path,
    detect_macro_source_from_path,
    is_macro_path,
)
from ingestion.pipeline import ingest_file, ingest_fred_series, ingest_macro_file
from storage.repositories import get_macro_series
from tests.test_screener_adapter import _make_screener_workbook


@pytest.fixture
def db_conn_with_companies(db_conn: sqlite3.Connection) -> sqlite3.Connection:
    seed_companies(db_conn)
    return db_conn


def test_ingest_file_inserts_observations_and_reconciles(
    tmp_path: Path, db_conn_with_companies: sqlite3.Connection
) -> None:
    file_path = tmp_path / "HDFCBANK.xlsx"
    _make_screener_workbook(file_path)

    result = ingest_file(db_conn_with_companies, file_path, company_id="HDFCBANK", source_id="screener")

    assert result.parsed_count > 0
    assert result.inserted_count == result.parsed_count
    assert result.skipped_count == 0
    assert result.reconciled_count == result.inserted_count  # every key is a fresh, distinct key

    row = db_conn_with_companies.execute(
        """
        SELECT canonical_value, reconciliation_reason FROM canonical_financials
        WHERE company_id = 'HDFCBANK' AND metric_key = 'net_profit'
          AND period_type = 'annual' AND fiscal_year = 'FY2024'
        """
    ).fetchone()
    assert row["canonical_value"] == 20500.0
    assert row["reconciliation_reason"] == "only source available"

    log_rows = db_conn_with_companies.execute(
        "SELECT was_chosen FROM reconciliation_log"
    ).fetchall()
    assert all(r["was_chosen"] == 1 for r in log_rows)  # single source -> every candidate is "the" candidate


def test_ingest_file_gate_blocks_archived_company(
    tmp_path: Path, db_conn_with_companies: sqlite3.Connection
) -> None:
    from companies.lifecycle import CompanyNotActiveError

    archive_company(db_conn_with_companies, "HDFCBANK", "manual")
    file_path = tmp_path / "HDFCBANK.xlsx"
    _make_screener_workbook(file_path)

    with pytest.raises(CompanyNotActiveError):
        ingest_file(db_conn_with_companies, file_path, company_id="HDFCBANK", source_id="screener")


def test_reingesting_same_file_supersedes_canonical_value(
    tmp_path: Path, db_conn_with_companies: sqlite3.Connection
) -> None:
    file_path = tmp_path / "HDFCBANK.xlsx"
    _make_screener_workbook(file_path)

    ingest_file(db_conn_with_companies, file_path, company_id="HDFCBANK", source_id="screener")
    second = ingest_file(db_conn_with_companies, file_path, company_id="HDFCBANK", source_id="screener")

    # Second ingestion adds new observation rows (raw data is never overwritten)...
    assert second.inserted_count > 0
    obs_count = db_conn_with_companies.execute(
        "SELECT COUNT(*) AS n FROM financial_observations WHERE company_id = 'HDFCBANK' AND metric_key = 'net_profit'"
    ).fetchone()["n"]
    # Fixture has 2 annual (FY2023, FY2024) + 4 quarterly (Q1-Q4 FY2024) net_profit
    # rows per ingestion = 6; doubled by re-ingestion since raw rows are never overwritten.
    assert obs_count == 12

    # ...but canonical_financials still resolves to exactly one row per key, pointing at the latest.
    canonical = db_conn_with_companies.execute(
        """
        SELECT chosen_observation_id FROM canonical_financials
        WHERE company_id = 'HDFCBANK' AND metric_key = 'net_profit'
          AND period_type = 'annual' AND fiscal_year = 'FY2024'
        """
    ).fetchone()
    latest_observation_id = db_conn_with_companies.execute(
        """
        SELECT MAX(observation_id) AS id FROM financial_observations
        WHERE company_id = 'HDFCBANK' AND metric_key = 'net_profit'
          AND period_type = 'annual' AND fiscal_year = 'FY2024'
        """
    ).fetchone()["id"]
    assert canonical["chosen_observation_id"] == latest_observation_id


def test_reconcile_updates_annual_canonical_row_in_place(
    tmp_path: Path, db_conn_with_companies: sqlite3.Connection
) -> None:
    """Regression test: SQLite's UNIQUE/ON CONFLICT never fires when a NULL
    participates in the unique columns (quarter is NULL for annual metrics),
    so a naive "INSERT ... ON CONFLICT DO UPDATE" silently duplicates the
    canonical row on every re-run instead of updating it in place.
    """
    file_path = tmp_path / "HDFCBANK.xlsx"
    _make_screener_workbook(file_path)

    ingest_file(db_conn_with_companies, file_path, company_id="HDFCBANK", source_id="screener")
    ingest_file(db_conn_with_companies, file_path, company_id="HDFCBANK", source_id="screener")

    rows = db_conn_with_companies.execute(
        """
        SELECT canonical_id FROM canonical_financials
        WHERE company_id = 'HDFCBANK' AND metric_key = 'net_profit'
          AND period_type = 'annual' AND fiscal_year = 'FY2024'
        """
    ).fetchall()
    assert len(rows) == 1


def test_detect_from_path_infers_company_and_source(tmp_path: Path) -> None:
    raw_dir = tmp_path / "data" / "raw"
    file_path = raw_dir / "HDFCBANK" / "screener" / "HDFCBANK.xlsx"
    file_path.parent.mkdir(parents=True)
    file_path.touch()

    company_id, source_id = detect_from_path(file_path, raw_dir=raw_dir)
    assert company_id == "HDFCBANK"
    assert source_id == "screener"


def test_detect_from_path_rejects_unknown_source(tmp_path: Path) -> None:
    raw_dir = tmp_path / "data" / "raw"
    file_path = raw_dir / "HDFCBANK" / "some_unregistered_source" / "file.xlsx"
    file_path.parent.mkdir(parents=True)
    file_path.touch()

    with pytest.raises(PathConventionError):
        detect_from_path(file_path, raw_dir=raw_dir)


def test_detect_from_path_rejects_path_outside_raw_dir(tmp_path: Path) -> None:
    raw_dir = tmp_path / "data" / "raw"
    raw_dir.mkdir(parents=True)
    outside_file = tmp_path / "elsewhere.xlsx"
    outside_file.touch()

    with pytest.raises(PathConventionError):
        detect_from_path(outside_file, raw_dir=raw_dir)


def test_ingest_file_normalizes_company_id_from_path(
    tmp_path: Path, db_conn_with_companies: sqlite3.Connection
) -> None:
    """Regression test: a raw.raw/<company> folder isn't necessarily typed in
    canonical case ("JioFinancial" rather than "JIOFINANCIAL"), but
    companies.company_id always is — path-detected company_id must be
    normalized the same way, or observations get inserted under a different
    company_id than the one they were registered under and analyze finds nothing.
    """
    from companies.registry import register_company

    register_company(
        db_conn_with_companies, "JIOFINANCIAL", "Jio Financial Services Ltd", "Jio Financial Services",
        sector="Financial Services", industry="NBFC",
    )

    raw_dir = tmp_path / "data" / "raw"
    file_path = raw_dir / "JioFinancial" / "screener" / "JioFinancial.xlsx"  # mixed-case folder
    file_path.parent.mkdir(parents=True)
    _make_screener_workbook(file_path)

    import config.settings as settings_module

    original_raw_dir = settings_module.RAW_DIR
    settings_module.RAW_DIR = raw_dir
    try:
        result = ingest_file(db_conn_with_companies, file_path)
    finally:
        settings_module.RAW_DIR = original_raw_dir

    assert result.company_id == "JIOFINANCIAL"
    row = db_conn_with_companies.execute(
        "SELECT COUNT(*) AS n FROM financial_observations WHERE company_id = 'JIOFINANCIAL'"
    ).fetchone()
    assert row["n"] == result.inserted_count > 0


def test_ingest_file_via_path_convention(tmp_path: Path, db_conn_with_companies: sqlite3.Connection) -> None:
    raw_dir = tmp_path / "data" / "raw"
    file_path = raw_dir / "HDFCBANK" / "screener" / "HDFCBANK.xlsx"
    file_path.parent.mkdir(parents=True)
    _make_screener_workbook(file_path)

    import config.settings as settings_module

    original_raw_dir = settings_module.RAW_DIR
    settings_module.RAW_DIR = raw_dir
    try:
        result = ingest_file(db_conn_with_companies, file_path)
    finally:
        settings_module.RAW_DIR = original_raw_dir

    assert result.company_id == "HDFCBANK"
    assert result.source_id == "screener"
    assert result.inserted_count > 0


# ------------------------------------------------------------------
# Macro (non-company) sources — data/raw/_macro/<source>/<file>
# ------------------------------------------------------------------


def test_is_macro_path_true_under_macro_sentinel(tmp_path: Path) -> None:
    raw_dir = tmp_path / "data" / "raw"
    file_path = raw_dir / "_macro" / "imd" / "rainfall_index.csv"
    assert is_macro_path(file_path, raw_dir=raw_dir) is True


def test_is_macro_path_false_for_company_folder(tmp_path: Path) -> None:
    raw_dir = tmp_path / "data" / "raw"
    file_path = raw_dir / "HDFCBANK" / "screener" / "HDFCBANK.xlsx"
    assert is_macro_path(file_path, raw_dir=raw_dir) is False


def test_detect_macro_source_from_path(tmp_path: Path) -> None:
    raw_dir = tmp_path / "data" / "raw"
    file_path = raw_dir / "_macro" / "imd" / "rainfall_index.csv"
    assert detect_macro_source_from_path(file_path, raw_dir=raw_dir) == "imd"


def test_detect_macro_source_from_path_rejects_non_macro_path(tmp_path: Path) -> None:
    raw_dir = tmp_path / "data" / "raw"
    file_path = raw_dir / "HDFCBANK" / "screener" / "HDFCBANK.xlsx"
    with pytest.raises(PathConventionError):
        detect_macro_source_from_path(file_path, raw_dir=raw_dir)


def test_detect_from_path_rejects_macro_path_with_a_clear_pointer(tmp_path: Path) -> None:
    """A macro path fed to the company-path detector should say so, not fail
    with a confusing "no adapter for source_id='imd'" (which would happen if
    it silently treated "_macro" as the company_id)."""
    raw_dir = tmp_path / "data" / "raw"
    file_path = raw_dir / "_macro" / "imd" / "rainfall_index.csv"
    with pytest.raises(PathConventionError, match="macro"):
        detect_from_path(file_path, raw_dir=raw_dir)


def test_ingest_macro_file_end_to_end(tmp_path: Path, db_conn: sqlite3.Connection) -> None:
    """No companies fixture needed — macro ingestion never touches the
    companies table (README: Data Layers -> Non-company sources)."""
    file_path = tmp_path / "rainfall_index.csv"
    file_path.write_text("period,value,unit\n2015,1108.9,MILLIMETRES\n2016,1142.3,MILLIMETRES\n")

    result = ingest_macro_file(db_conn, file_path, source_id="imd")

    assert result.series_key == "rainfall_index"
    assert result.source_id == "imd"
    assert result.parsed_count == 2
    assert result.inserted_count == 2
    assert result.skipped_count == 0

    series = get_macro_series(db_conn, "rainfall_index")
    assert [row["period"] for row in series] == ["2015", "2016"]
    assert [row["value"] for row in series] == [1108.9, 1142.3]
    assert all(row["source"] == "imd" for row in series)
    assert all(row["region"] is None for row in series)


def test_ingest_fred_series_end_to_end(monkeypatch, db_conn: sqlite3.Connection) -> None:
    """No companies fixture needed, same as macro ingestion above — FRED
    series aren't scoped to a company either. HTTP is mocked, same as
    tests/test_fred_adapter.py."""
    from contextlib import contextmanager

    class _FakeResponse:
        def __init__(self, body: bytes) -> None:
            self._body = body

        def read(self) -> bytes:
            return self._body

    @contextmanager
    def fake_urlopen(req, timeout=None):
        yield _FakeResponse(b"observation_date,FEDFUNDS\n2020-01-01,1.55\n2020-02-01,1.58\n")

    monkeypatch.setattr("sources.fred.urllib.request.urlopen", fake_urlopen)

    result = ingest_fred_series(db_conn, "FEDFUNDS", unit="PERCENT")

    assert result.source_id == "fred"
    assert result.series_key == "fedfunds"
    assert result.parsed_count == 2
    assert result.inserted_count == 2
    assert result.skipped_count == 0

    series = get_macro_series(db_conn, "fedfunds")
    assert [row["value"] for row in series] == [1.55, 1.58]
    assert all(row["source"] == "fred" for row in series)


def test_ingest_macro_file_skips_invalid_rows_with_reasons(tmp_path: Path, db_conn: sqlite3.Connection) -> None:
    file_path = tmp_path / "rainfall_index.csv"
    file_path.write_text("period,value,unit\n2015,1108.9,MILLIMETRES\nFY2016,100,MILLIMETRES\n")

    result = ingest_macro_file(db_conn, file_path, source_id="imd")

    assert result.parsed_count == 1  # the adapter itself already drops the malformed period row
    assert result.inserted_count == 1
    assert result.skipped_count == 0


def test_ingest_macro_file_routes_xlsx_to_the_rbi_indicator_workbook_parser(
    tmp_path: Path, db_conn: sqlite3.Connection
) -> None:
    import openpyxl

    from ingestion.pipeline import ingest_macro_file

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in ("Weekly", "Fortnightly", "Monthly", "Quarterly"):
        ws = wb.create_sheet(sheet_name)
        ws.append([None] * 2)
        ws.append([None, f"Macro-economic Indicators - {sheet_name}"])
        ws.append([None])
        ws.append([None, "Period", "Policy Repo Rate (%)"])
    from datetime import datetime
    wb["Weekly"].append([None, datetime(2026, 8, 7), 5.25])
    wb["Monthly"].append([None, "Jul-2026", 5.25])
    file_path = tmp_path / "50 Macroeconomic Indicators.xlsx"
    wb.save(file_path)

    result = ingest_macro_file(db_conn, file_path, source_id="rbi")

    assert result.inserted_count >= 1
    series = get_macro_series(db_conn, "policy_repo_rate")
    assert any(row["period"] == "2026-08-07" for row in series)


def test_ingest_macro_file_routes_other_xlsx_to_the_dbie_table_parser(
    tmp_path: Path, db_conn: sqlite3.Connection
) -> None:
    import openpyxl

    from ingestion.pipeline import ingest_macro_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None] * 4)
    ws.append([None, "No. 21: Monthly Average Price of Gold and Silver in Mumbai"])
    ws.append([None, "Item", "2025-26", 2026])
    ws.append([None, None, None, "May"])
    ws.append([None, None, 1, 2])
    ws.append([None, "1. Standard Gold (₹ per 10 grams)", 118421, 154756])
    file_path = tmp_path / "21T.XLSX"
    wb.save(file_path)

    result = ingest_macro_file(db_conn, file_path, source_id="rbi")

    # Column D ("2026" + "May") resolves directly; column C ("2025-26", a
    # bare fiscal-year label with no month/day of its own) backfills from
    # D's resolved month/day, giving both real RBI-file dates -- confirmed
    # against the actual table 21 export during development.
    series = get_macro_series(db_conn, "t21_1_standard_gold_per_10_grams")
    assert {row["period"] for row in series} == {"2026-05-31", "2025-05-31"}
    assert all(row["period_type"] == "dated" for row in series)


def test_ingest_macro_file_series_key_override(tmp_path: Path, db_conn: sqlite3.Connection) -> None:
    file_path = tmp_path / "data.csv"
    file_path.write_text("period,value,unit\n2015,6.75,PERCENT\n")

    result = ingest_macro_file(db_conn, file_path, source_id="rbi", series_key="repo_rate")

    assert result.series_key == "repo_rate"
    series = get_macro_series(db_conn, "repo_rate")
    assert len(series) == 1


def test_ingest_macro_file_infers_source_from_path(tmp_path: Path, db_conn: sqlite3.Connection) -> None:
    raw_dir = tmp_path / "data" / "raw"
    file_path = raw_dir / "_macro" / "imd" / "rainfall_index.csv"
    file_path.parent.mkdir(parents=True)
    file_path.write_text("period,value,unit\n2015,1108.9,MILLIMETRES\n")

    import config.settings as settings_module

    original_raw_dir = settings_module.RAW_DIR
    settings_module.RAW_DIR = raw_dir
    try:
        result = ingest_macro_file(db_conn, file_path)
    finally:
        settings_module.RAW_DIR = original_raw_dir

    assert result.source_id == "imd"
    assert result.series_key == "rainfall_index"


def test_get_macro_series_distinguishes_region(tmp_path: Path, db_conn: sqlite3.Connection) -> None:
    file_path = tmp_path / "rainfall.csv"
    file_path.write_text(
        "period,value,unit,region\n2015,1108.9,MILLIMETRES,\n2015,950.2,MILLIMETRES,Maharashtra\n"
    )
    ingest_macro_file(db_conn, file_path, source_id="imd", series_key="rainfall")

    all_india = get_macro_series(db_conn, "rainfall")
    maharashtra = get_macro_series(db_conn, "rainfall", region="Maharashtra")

    assert [row["value"] for row in all_india] == [1108.9]
    assert [row["value"] for row in maharashtra] == [950.2]
