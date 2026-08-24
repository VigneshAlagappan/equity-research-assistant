"""sources/rbi_indicators.py — parsing the RBI "50 Macroeconomic Indicators"
workbook. Tests build a small synthetic workbook with the same shape as the
real publication rather than depending on the real downloaded file."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from sources.rbi_indicators import looks_like_rbi_indicator_workbook, parse_rbi_indicator_workbook


def _make_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    weekly = wb.create_sheet("Weekly")
    weekly.append([None] * 2)
    weekly.append([None, "Macro-economic Indicators - Weekly"])
    weekly.append([None])
    weekly.append([None, "Period", "Policy Repo Rate (%)", "Foreign Exchange Reserves \n(US $ Million)"])
    weekly.append([None, datetime(2026, 8, 7), 5.25, 707001.696])
    weekly.append([None, datetime(2026, 7, 31), 5.25, 692865.992])

    monthly = wb.create_sheet("Monthly")
    monthly.append([None] * 2)
    monthly.append([None, "Macro-economic Indicators - Monthly"])
    monthly.append([None])
    monthly.append([None, "Period", "Consumer Price Index  (2012=100)", "Consumer Price Index  (2024=100)"])
    monthly.append([None, "Jul-2026", "-", 107.94])
    monthly.append([None, "Jun-2026", "-", 107.0])

    fortnightly = wb.create_sheet("Fortnightly")
    fortnightly.append([None] * 2)
    fortnightly.append([None, "Macro-economic Indicators - Fortnightly"])
    fortnightly.append([None])
    fortnightly.append([None, "Period", "M3 \n(₹ Crore)"])
    fortnightly.append([None, datetime(2026, 7, 31), 32281301.87])

    quarterly = wb.create_sheet("Quarterly")
    quarterly.append([None] * 2)
    quarterly.append([None, "Macro-economic Indicators - Quarterly"])
    quarterly.append([None])
    quarterly.append([None, "Period", "International Investment Position Net\n(US $ Million)"])
    quarterly.append([None, datetime(2026, 3, 31), -209950.301930149])

    wb.save(path)


@pytest.fixture
def workbook_path(tmp_path: Path) -> Path:
    path = tmp_path / "50 Macroeconomic Indicators.xlsx"
    _make_workbook(path)
    return path


def test_looks_like_rbi_indicator_workbook_true_for_matching_sheets(workbook_path: Path) -> None:
    assert looks_like_rbi_indicator_workbook(workbook_path) is True


def test_looks_like_rbi_indicator_workbook_false_for_other_files(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    path = tmp_path / "not_it.xlsx"
    wb.save(path)
    assert looks_like_rbi_indicator_workbook(path) is False


def test_parses_weekly_sheet_with_real_dates(workbook_path: Path) -> None:
    obs = parse_rbi_indicator_workbook(workbook_path)
    repo_rate = [o for o in obs if o.series_key == "policy_repo_rate"]
    assert len(repo_rate) == 2
    assert {o.period for o in repo_rate} == {"2026-08-07", "2026-07-31"}
    assert all(o.period_type == "weekly" for o in repo_rate)
    assert all(o.unit == "PERCENT" for o in repo_rate)


def test_parses_monthly_sheet_from_mon_yyyy_strings(workbook_path: Path) -> None:
    obs = parse_rbi_indicator_workbook(workbook_path)
    cpi_2024 = [o for o in obs if o.series_key == "consumer_price_index_2024_100"]
    assert {o.period for o in cpi_2024} == {"2026-07", "2026-06"}
    assert all(o.period_type == "monthly" for o in cpi_2024)


def test_different_cpi_base_years_do_not_collide(workbook_path: Path) -> None:
    """Regression: "(2012=100)" and "(2024=100)" must not collapse to the
    same series_key just because both are parenthetical."""
    obs = parse_rbi_indicator_workbook(workbook_path)
    series = {o.series_key for o in obs}
    assert "consumer_price_index_2012_100" not in series  # every value was "-" (skipped)
    assert "consumer_price_index_2024_100" in series


def test_dash_placeholder_values_are_skipped(workbook_path: Path) -> None:
    obs = parse_rbi_indicator_workbook(workbook_path)
    assert all(o.series_key != "consumer_price_index_2012_100" for o in obs)  # both rows were "-"


def test_unit_inference_from_header_hints(workbook_path: Path) -> None:
    obs = parse_rbi_indicator_workbook(workbook_path)
    forex = next(o for o in obs if o.series_key == "foreign_exchange_reserves")
    assert forex.unit == "USD_MILLION"
    m3 = next(o for o in obs if o.series_key == "m3")
    assert m3.unit == "INR_CRORE"


def test_fortnightly_and_quarterly_use_dated_period_shape(workbook_path: Path) -> None:
    obs = parse_rbi_indicator_workbook(workbook_path)
    m3 = next(o for o in obs if o.series_key == "m3")
    assert m3.period_type == "fortnightly"
    assert m3.period == "2026-07-31"
    iip_net = next(o for o in obs if "international_investment_position" in o.series_key)
    assert iip_net.period_type == "quarterly"
    assert iip_net.period == "2026-03-31"


def test_raises_for_a_workbook_missing_expected_sheets(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    path = tmp_path / "unrelated.xlsx"
    wb.save(path)
    with pytest.raises(ValueError):
        parse_rbi_indicator_workbook(path)
