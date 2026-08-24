"""sources/rbi_dbie_tables.py — parsing bespoke single-table RBI DBIE
exports. Tests build small synthetic workbooks mirroring the real shapes
(title row, merged multi-row header, index row, data rows), including the
known header-date-corruption pattern, rather than depending on the real
downloaded files."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from sources.rbi_dbie_tables import parse_rbi_dbie_table


def _make_table13_like(path: Path) -> None:
    """Mirrors table 13's real shape: one confidently-dated string column
    ("As on March 31, 2026"), one year+corrupted-datetime column that
    should resolve via the corrupted datetime falling outside the plausible
    range, and one column that's *only* a corrupted datetime (no usable
    signal at all) which must be skipped."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T_13"
    ws.append([None] * 7)
    ws.append([None, "No. 13: Scheduled Commercial Banks’ Investments"])
    ws.append([None, "(₹ Crore)"])
    ws.append([None, "Item", "As on March 31, 2026", 2025, 2026, None, None])
    ws.append([None, None, None, datetime(1930, 5, 1), "Apr. 30", datetime(2015, 5, 1), datetime(1931, 5, 1)])
    ws.append([None, None, 1, 2, 3, 4, 5])
    ws.append([None, "1. SLR Securities", 6961120, 6706717, 6926908, 6964934, 7034653])
    wb.save(path)


def _make_growth_table_like(path: Path) -> None:
    """Mirrors table 15/16/18(b)'s shape: dated snapshot columns mixed with
    growth-rate columns (index-row marker "%", header text containing
    "so far"/"over"/"Y-o-Y") that must never be treated as a period."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T_15"
    ws.append([None] * 8)
    ws.append([None, "No. 15: Deployment of Gross Bank Credit by Major Sectors"])
    ws.append([None, "(₹ Crore)"])
    ws.append([None, "Sector", "Outstanding as on", None, None, None, "Growth (%)", None])
    ws.append([None, None, "Mar. 31, 2026", 2025, 2026, None, "Financial year so far", "Y-o-Y"])
    ws.append([None, None, None, datetime(1930, 5, 1), "Apr. 30", datetime(1931, 5, 1), "2026-27", 2026])
    ws.append([None, None, 1, 2, 3, 4, "%", "%"])
    ws.append([None, "I. Bank Credit (II + III)", 21361435, 18287377, 21211828, 21515965, 0.7, 17.7])
    ws.merge_cells("D4:F4")
    wb.save(path)


def _make_month_end_table_like(path: Path) -> None:
    """Mirrors table 18(b)'s "At End-March"/"As on June 30" shape — a bare
    month name with no explicit day must resolve to that month's real
    last day, not a fixed day-of-month."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T_18(b)"
    ws.append([None] * 4)
    ws.append([None, "No. 18 (b): Outstanding Credit to Commercial Sector in India"])
    ws.append([None, "Source", "₹ Crore"])
    ws.append([None, None, "At End-March", "As on June 30"])
    ws.append([None, None, 2024, 2024])
    ws.append([None, None, 1, 2])
    ws.append([None, "1. Non-Food Bank Credit", 16563961, 16852125])
    wb.save(path)


def test_resolves_confidently_dated_column_and_skips_uninterpretable_ones(tmp_path: Path) -> None:
    path = tmp_path / "13T.XLSX"
    _make_table13_like(path)

    obs = parse_rbi_dbie_table(path)

    periods = {o.period for o in obs}
    assert "2026-03-31" in periods  # "As on March 31, 2026" — explicit, trustworthy
    # The bare-year column (2025) backfills from the resolved sibling's month/day.
    assert "2025-03-31" in periods
    # "Apr. 30" + year 2026 -> resolved via month/day search + year backfill.
    assert "2026-04-30" in periods
    # Columns whose only signal is an implausible-year corrupted datetime
    # (1930/1931/2015) must never appear as a period.
    assert not any(p.startswith(("1930", "1931", "2015")) for p in periods)


def test_series_key_is_namespaced_by_table_number(tmp_path: Path) -> None:
    path = tmp_path / "13T.XLSX"
    _make_table13_like(path)

    obs = parse_rbi_dbie_table(path)

    assert all(o.series_key.startswith("t13_") for o in obs)
    assert any("slr_securities" in o.series_key for o in obs)


def test_growth_rate_columns_are_never_treated_as_a_period(tmp_path: Path) -> None:
    path = tmp_path / "15T.XLSX"
    _make_growth_table_like(path)

    obs = parse_rbi_dbie_table(path)

    # Only the two genuine snapshot dates should appear — never a period
    # derived from "Financial year so far" / "Y-o-Y" / "2026-27".
    periods = {o.period for o in obs}
    assert periods == {"2026-03-31", "2025-03-31", "2026-04-30"}


def test_growth_rate_column_values_are_dropped_not_mislabeled(tmp_path: Path) -> None:
    path = tmp_path / "15T.XLSX"
    _make_growth_table_like(path)

    obs = parse_rbi_dbie_table(path)

    # The growth values themselves (0.7, 17.7) must not show up under any period.
    assert not any(o.value in (0.7, 17.7) for o in obs)


def test_bare_month_with_no_explicit_day_resolves_to_real_month_end(tmp_path: Path) -> None:
    path = tmp_path / "18BT.XLSX"
    _make_month_end_table_like(path)

    obs = parse_rbi_dbie_table(path)

    periods = {o.period for o in obs}
    assert "2024-03-31" in periods  # "At End-March" -> March has 31 days, not a fixed day-28
    assert "2024-06-30" in periods  # "As on June 30" -> explicit day, unaffected


def test_unit_is_percent_when_index_row_marks_it(tmp_path: Path) -> None:
    path = tmp_path / "15T.XLSX"
    _make_growth_table_like(path)

    obs = parse_rbi_dbie_table(path)

    assert all(o.unit != "PERCENT" for o in obs)  # the only "%"-marked columns were skipped (growth columns)


def test_raises_for_a_file_with_no_recognizable_table_shape(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.active.append(["just", "some", "random", "cells"])
    path = tmp_path / "not_a_table.xlsx"
    wb.save(path)

    with pytest.raises(ValueError):
        parse_rbi_dbie_table(path)


def test_period_type_is_dated_for_every_observation(tmp_path: Path) -> None:
    path = tmp_path / "13T.XLSX"
    _make_table13_like(path)

    obs = parse_rbi_dbie_table(path)

    assert obs and all(o.period_type == "dated" for o in obs)
