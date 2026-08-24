"""Parses an "Equity Analysis" workbook's Forecast sheet (the same personal
template used for `data/templates/EquityAnalysis-Bank-HDFCBank.xlsx`, whose
hand-ported values in `web/static/data/hdfcbank-analysis.json` were used as
ground truth to build the ROW_LABELS table below — every (section, key)
value was verified against its source row, byte-for-byte) into the
{"YEARS": [...], "METRICS": {...}} shape the valuation dashboard consumes.

Not hand-verified per new company the way HDFC's file was — this is an
automated best-effort parse against a template whose row layout is
consistent across files but whose *filled-in* rows vary (a company template
built around a non-bank NBFC, for instance, may fill different candidate
rows for "investments" than a bank one does). Spot-check the dashboard for
any company this produces before trusting a specific number from it.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

# (section, key, dashboard_label, unit, source_row_label). Where a metric
# could plausibly live under more than one row across different company
# templates (e.g. a bank's "Investments" bank-book row vs. a generic
# current-assets "Investments" row), source_row_label is the one HDFC's
# template actually used — see module docstring.
ROW_LABELS: list[tuple[str, str, str, str, str]] = [
    ("balanceSheet", "networth", "Networth (reserves only)", "crore", "Networth (reserves only)"),
    ("balanceSheet", "she", "Shareholders Equity (SHE)", "crore", "SHE (includes reserves & Share capital)"),
    ("balanceSheet", "deposits", "Gross Deposits", "crore", "Gross Deposits"),
    ("balanceSheet", "borrowings", "Gross Borrowings", "crore", "Gross Borrowings"),
    ("balanceSheet", "advances", "Advances", "crore", "Advances"),
    ("balanceSheet", "investments", "Investments", "crore", "Investments"),
    ("balanceSheet", "totalAssets", "Total Assets / Liabilities", "crore", "Total Assets/Liabilities"),
    ("incomeStatement", "earnings", "Earnings (Total Income)", "crore", "Earnings"),
    ("incomeStatement", "expenses", "Expenses", "crore", "Expenses"),
    ("incomeStatement", "interestOutgo", "Interest Out-go", "crore", "Interest Out-go"),
    ("incomeStatement", "otherIncome", "Other Income", "crore", "Other Income"),
    ("incomeStatement", "depreciation", "Depreciation", "crore", "Depreciation"),
    ("incomeStatement", "netProfit", "Net Profit (PAT)", "crore", "Net Earnings (Net Profit or PAT)"),
    ("perShare", "eps", "EPS (Net Profit / share)", "rupee", "EPS (Net Profit per share)"),
    ("perShare", "bookValue", "Book Value (Networth based)", "rupee", "Book Value (Networth based) per share"),
    ("perShare", "dividend", "Dividend per share", "rupee", "Dividend (per share)"),
    ("perShare", "salesPerShare", "Sales (Revenue per share)", "rupee", "Sales (Rev per share)"),
    ("perShare", "shares", "Shares Outstanding", "crShares", "Total Number of outstanding shares (in Cr)"),
    ("profitability", "netMargin", "Net Profit Margin", "pct", "Net Profit Margin"),
    ("profitability", "roe", "RONW / ROE", "pct", "RONW or ROE"),
    ("profitability", "taxRate", "Tax Paid % of Revenue", "pctAbs", "Tax Paid % (Tax Paid/ Revenue)"),
    ("profitability", "payout", "Dividend Payout Ratio", "pct", "Dividend/Net Earnings ratio"),
    ("profitability", "retention", "Retention Ratio", "pct", "Retention Ratio"),
    ("bankRatios", "cdRatio", "Credit-Deposit Ratio", "pct", "Credit Deposit Ratio (advances / Deposits)"),
    ("bankRatios", "advDepBorrow", "Advances / (Deposits + Borrowings)", "pct", "Advances / (gross Deposit + Borrowings)"),
    ("bankRatios", "otherIncEarn", "Other Income / Earnings", "pct", "Other Income / Earnings"),
    ("bankRatios", "npAssets", "Net Profit / Total Assets", "pct", "Net Profit / (Total Assets/Liabilities)"),
    ("bankRatios", "intCoverage", "Interest Coverage (Earnings / Interest Outgo)", "x", "Interest Coverage ratio (Earnings/Interest Outgo)"),
    ("bankRatios", "intOverProfit", "Interest Outgo / Net Profit", "x", "Interest Out go/ Net Profit"),
    ("valuation", "price", "Price", "rupee", "Price"),
    ("valuation", "pe", "P/E", "x", "P/E (how many years it will take to recover invested amount)"),
    ("valuation", "pbv", "P/BV (Shareholder Equity)", "x", "P/BV (Share Holder Equity)"),
    ("valuation", "divYield", "Dividend Yield", "pct", "Dividend Yield"),
]

_SECTION_ORDER = ["balanceSheet", "incomeStatement", "perShare", "profitability", "bankRatios", "valuation"]

# Rows whose union of non-null years defines the company's overall YEARS
# range — the core actuals, not every derived ratio (a ratio can be null in
# a year purely because its inputs divide-by-zero, which shouldn't trim an
# otherwise-populated year).
_ANCHOR_LABELS = {
    "Networth (reserves only)", "SHE (includes reserves & Share capital)", "Gross Deposits", "Gross Borrowings",
    "Advances", "Investments", "Total Assets/Liabilities", "Earnings", "Expenses", "Interest Out-go",
    "Other Income", "Depreciation", "Net Earnings (Net Profit or PAT)",
}


def find_forecast_sheet(workbook: openpyxl.Workbook):
    for name in workbook.sheetnames:
        if "forecast" in name.lower():
            return workbook[name]
    raise ValueError(f"No sheet with 'Forecast' in its name: {workbook.sheetnames}")


def _clean(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None  # blank cells, error strings ("#DIV/0!"), text — all treated as no data


def _is_year(value: Any) -> bool:
    return isinstance(value, (int, float)) and 1990 <= value <= 2100


def parse_forecast_sheet(worksheet) -> dict:
    year_header_row = None
    for row in worksheet.iter_rows(min_row=1, max_row=10):
        if sum(1 for cell in row if _is_year(cell.value)) >= 3:
            year_header_row = row
            break
    if year_header_row is None:
        raise ValueError("Could not find the year-header row (expected several 19xx/20xx values)")

    # Only the first contiguous run of year cells — the template appends a
    # second, separately-columned "projected forward" year block later in
    # the same row (e.g. 2024-2033 after a "CAGR"/"Proj CAGR" text gap),
    # which is assumption-driven forward projection, not recorded actuals;
    # excluded on the same "facts only" basis as everywhere else in this app.
    year_col_start = next(i for i, cell in enumerate(year_header_row) if _is_year(cell.value))
    all_years: list[int] = []
    for cell in year_header_row[year_col_start:]:
        if not _is_year(cell.value):
            break
        all_years.append(cell.value)

    label_to_values: dict[str, list[float | None]] = {}
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        label = row[1].value if len(row) > 1 else None
        if not isinstance(label, str) or not label.strip():
            continue
        label = label.strip()
        values = [_clean(cell.value) for cell in row[year_col_start : year_col_start + len(all_years)]]
        if label not in label_to_values or not any(v is not None for v in label_to_values[label]):
            label_to_values[label] = values

    anchor_years = [
        year
        for i, year in enumerate(all_years)
        if any((label_to_values.get(anchor) or [None] * len(all_years))[i] is not None for anchor in _ANCHOR_LABELS)
    ]
    if not anchor_years:
        years: list[int] = []
    else:
        start, end = min(anchor_years), max(anchor_years)
        years = [y for y in all_years if start <= y <= end]

    year_indices = [all_years.index(y) for y in years]

    metrics: dict[str, list[dict]] = {section: [] for section in _SECTION_ORDER}
    for section, key, label, unit, source_label in ROW_LABELS:
        raw_values = label_to_values.get(source_label, [None] * len(all_years))
        trimmed = [raw_values[i] for i in year_indices]
        metrics[section].append({"key": key, "label": label, "unit": unit, "values": trimmed})

    return {"YEARS": [int(y) for y in years], "METRICS": metrics}


def parse_workbook(path: Path) -> dict:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = find_forecast_sheet(workbook)
    return parse_forecast_sheet(sheet)
