"""ProprietaryAdapter — parses a hand-built "Equity Analysis" workbook's own
"3 - Forecast" sheet, e.g. data/raw/ARE&M/proprietary/..._Amara_Raja_
Battries.xlsx (uploaded via Admin > Import Data, source=proprietary).

A completely different layout from Screener's Data Sheet tab (sources/
screener.py) — this is a personal valuation-model template, not a vendor
export. "3 - Forecast" is one wide grid: a "Company Name" row, immediately
followed by a year-header row (plain 4-digit years — 2000, 2001, ... —
across columns starting at column C), then metric rows aligned to those
same columns, label in column B. The same header row's actual-year block
ends where a text cell ("CAGR", "Proj  CAGR", ...) breaks the run of years —
everything from there on is the model's own forward projection, not an
observed fact, so it's never read.

Row-label -> metric_key mapping goes through metric_aliases like every other
adapter (normalization/financials.py's DEFAULT_METRIC_ALIASES,
source="proprietary") — this module never hardcodes a metric name, only the
grid's shape.
"""

from __future__ import annotations

from storage.db_types import DBConnection
from pathlib import Path

import openpyxl

from normalization.financials import build_observations_from_periods
from sources.base import NormalizedObservation, SourceAdapter

PARSER_VERSION = "proprietary-v1-forecast-sheet"

FORECAST_SHEET_NAME = "3 - Forecast"
_COMPANY_NAME_LABEL = "COMPANY NAME"
_LABEL_COLUMN = 1  # column B (0-indexed)
_FIRST_YEAR_COLUMN = 2  # column C onward
_YEAR_RANGE = range(1980, 2100)


class ProprietaryAdapter(SourceAdapter):
    source_id = "proprietary"

    def __init__(self, conn: DBConnection):
        self._conn = conn

    def parse(
        self,
        file_path: Path,
        company_id: str,
        statement_type: str = "consolidated",
        **kwargs: object,
    ) -> list[NormalizedObservation]:
        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        try:
            if FORECAST_SHEET_NAME not in workbook.sheetnames:
                raise ValueError(
                    f"{file_path} has no {FORECAST_SHEET_NAME!r} tab — is this one of the "
                    f"hand-built Equity Analysis workbooks?"
                )
            rows = list(workbook[FORECAST_SHEET_NAME].iter_rows(values_only=True))
        finally:
            workbook.close()

        return self._parse_rows(rows, company_id=company_id, source_file=str(file_path), statement_type=statement_type)

    def _parse_rows(
        self, rows: list[tuple[object, ...]], *, company_id: str, source_file: str, statement_type: str
    ) -> list[NormalizedObservation]:
        header_row_index = self._find_year_header_row(rows)
        if header_row_index is None:
            raise ValueError(
                f"No year-header row found in {FORECAST_SHEET_NAME!r} — expected a 'Company Name' "
                f"row immediately followed by a row of 4-digit years"
            )
        year_columns = self._actual_year_columns(rows[header_row_index])

        observations: list[NormalizedObservation] = []
        for row in rows[header_row_index + 1 :]:
            label = row[_LABEL_COLUMN] if len(row) > _LABEL_COLUMN else None
            if not isinstance(label, str) or not label.strip():
                continue  # blank row, or a label in some other column (section markers, notes)

            period_values = {
                (f"FY{year}", None): row[col_index]
                for col_index, year in year_columns
                if col_index < len(row)
            }
            observations.extend(
                build_observations_from_periods(
                    self._conn,
                    company_id=company_id,
                    source=self.source_id,
                    source_file=source_file,
                    parser_version=PARSER_VERSION,
                    period_type="annual",
                    statement_type=statement_type,
                    row_label=label.strip(),
                    period_values=period_values,
                )
            )
        return observations

    @staticmethod
    def _find_year_header_row(rows: list[tuple[object, ...]]) -> int | None:
        for i, row in enumerate(rows):
            label = row[_LABEL_COLUMN] if len(row) > _LABEL_COLUMN else None
            if isinstance(label, str) and label.strip().upper() == _COMPANY_NAME_LABEL:
                return i + 1
        return None

    @staticmethod
    def _actual_year_columns(header_row: tuple[object, ...]) -> list[tuple[int, int]]:
        """Column index -> year, for the contiguous run of year cells
        starting at _FIRST_YEAR_COLUMN. Stops at the first non-year, non-blank
        cell ("CAGR", "Proj  CAGR", ...) — that boundary is where the
        model's own projected years begin, which this adapter never reads."""
        years: list[tuple[int, int]] = []
        for col_index in range(_FIRST_YEAR_COLUMN, len(header_row)):
            cell = header_row[col_index]
            if isinstance(cell, (int, float)) and int(cell) in _YEAR_RANGE:
                years.append((col_index, int(cell)))
            elif cell not in (None, ""):
                break
        return years
