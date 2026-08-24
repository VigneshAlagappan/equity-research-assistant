"""Parses RBI's monthly bank-level infrastructure/transaction bulletins —
"ATM, Acceptance Infrastructure and Card Statistics" and "NEFT"/"RTGS"
exports under data/raw/_macro/rbi/MoneyAndBanks/ (filenames starting
"ATM"/"NEFTRTGS"). Genuinely bank x metric x period data, not a flat
series x period the way every other RBI file in this app is — see
schemas/sqlite_schema.sql's bank_infrastructure_observations table.

Both file kinds share a shape: a title row naming the report month, several
header rows describing each column's metric (deeply nested for the ATM
file — up to 5 header rows; simpler for NEFT/RTGS — 2), then bank rows
(a serial-number column, a bank-name column, then one numeric value per
metric column). ATM additionally interleaves section-header rows
("Scheduled Commercial Banks", "Public Sector Banks") with no serial
number — skipped, not treated as a bank.

Unlike sources/rbi_dbie_tables.py's tables, there's no separate "index
row" of sequential 1,2,3... markers here reliably in every variant (NEFT
has none) — bank rows are instead recognized directly by their serial-
number column holding a positive integer.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl

PARSER_VERSION = "rbi-bank-infra-v1-xlsx"

_MONTH_NAMES = {m.lower(): i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1
)}
_TITLE_MONTH_YEAR_RE = re.compile(
    r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+(20\d{2})\b", re.IGNORECASE
)
_SERIAL_HEADER_RE = re.compile(r"^S[rl]\.?\s*No\.?$", re.IGNORECASE)
_BANK_NAME_HEADER_RE = re.compile(r"bank\s*name", re.IGNORECASE)

_UNIT_HINTS = [
    (re.compile(r"Rs\.?\s*Crore|₹\s*Crore", re.IGNORECASE), "INR_CRORE"),
    (re.compile(r"Rs'000|Rs\.\s*000", re.IGNORECASE), "INR_THOUSAND"),
]


@dataclass(frozen=True)
class BankInfrastructureObservation:
    bank_name: str
    metric: str
    period_type: str  # always "monthly"
    period: str  # "YYYY-MM"
    value: float
    unit: str
    source: str
    source_file: str
    parser_version: str


def _slugify(label: str) -> str:
    label = label.replace("–", "-").replace("’", "")
    label = re.sub(r"[^A-Za-z0-9]+", "_", label)
    return re.sub(r"_+", "_", label).strip("_").lower()


def _unit_for_label(label: str) -> str:
    for pattern, unit in _UNIT_HINTS:
        if pattern.search(label):
            return unit
    return "NUMBER"


def _merged_forward_fill(ws) -> dict[tuple[int, int], object]:
    fill: dict[tuple[int, int], object] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = ws.cell(merged_range.min_row, merged_range.min_col).value
        if top_left is None:
            continue
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                fill[(row, col)] = top_left
    return fill


def _cell_value(ws, merged: dict[tuple[int, int], object], row: int, col: int) -> object:
    value = ws.cell(row, col).value
    return value if value is not None else merged.get((row, col))


def _find_report_period(ws) -> str | None:
    """Month + year from the title row, e.g. "...for the Month of June
    2026" or "...(NEFT) - JUNE 2026" -> "2026-06"."""
    for row in ws.iter_rows(min_row=1, max_row=4):
        for cell in row:
            if isinstance(cell.value, str):
                match = _TITLE_MONTH_YEAR_RE.search(cell.value)
                if match:
                    month = _MONTH_NAMES[match.group(1)[:3].lower()]
                    return f"{match.group(2)}-{month:02d}"
    return None


def _find_serial_and_bank_columns(ws, merged: dict[tuple[int, int], object]) -> tuple[int, int, int] | None:
    """(serial_col, bank_col, last_header_row) by locating the "Bank Name"
    header cell — the serial column is assumed immediately to its left,
    matching both real file layouts."""
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if isinstance(cell.value, str) and _BANK_NAME_HEADER_RE.search(cell.value):
                return cell.column - 1, cell.column, cell.row
    return None


def _find_first_data_row(ws, serial_col: int, header_row: int) -> int | None:
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        value = row[serial_col - 1].value
        if isinstance(value, int) and not isinstance(value, bool) and value > 0:
            return row[0].row
    return None


def parse_bank_infrastructure_file(file_path: Path) -> list[BankInfrastructureObservation]:
    """Parse one ATM/NEFT/RTGS monthly bulletin. Raises ValueError if the
    file doesn't have a recognizable "Bank Name" column and report-month
    title — a defensive check, not a claim every metric column will be
    ingested (blank/non-numeric columns are silently skipped per row, same
    as every other RBI parser in this app)."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    period = _find_report_period(ws)
    columns = _find_serial_and_bank_columns(ws, {})
    if period is None or columns is None:
        raise ValueError(f"{file_path} does not look like an RBI bank-infrastructure bulletin")
    serial_col, bank_col, header_row = columns
    data_start_row = _find_first_data_row(ws, serial_col, header_row)
    if data_start_row is None:
        raise ValueError(f"{file_path}: found headers but no bank data rows")

    merged = _merged_forward_fill(ws)
    metric_cols: dict[int, str] = {}
    max_col = ws.max_column
    for col in range(bank_col + 1, max_col + 1):
        header_texts = [
            str(t) for r in range(header_row, data_start_row)
            if (t := _cell_value(ws, merged, r, col)) is not None and str(t).strip()
            # Excludes the bare column-index row (1, 2, 3...) some of these
            # bulletins carry right above the data — a small sequential
            # marker, not descriptive text; keeping it just clutters every
            # metric label with a trailing "_10", "_11", etc.
            and not (isinstance(t, int) and not isinstance(t, bool) and 0 < t <= 100)
        ]
        if header_texts:
            metric_cols[col] = " ".join(header_texts)

    observations: list[BankInfrastructureObservation] = []
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
        serial = row[serial_col - 1].value
        if not isinstance(serial, int) or isinstance(serial, bool) or serial <= 0:
            continue  # section-header row (e.g. "Public Sector Banks") or trailing footnote
        bank_name = row[bank_col - 1].value
        if not isinstance(bank_name, str) or not bank_name.strip():
            continue
        for col, label in metric_cols.items():
            value = row[col - 1].value
            if not isinstance(value, (int, float)) or isinstance(value, bool):
                continue
            observations.append(BankInfrastructureObservation(
                bank_name=bank_name.strip(), metric=_slugify(label), period_type="monthly", period=period,
                value=float(value), unit=_unit_for_label(label), source="rbi",
                source_file=str(file_path), parser_version=PARSER_VERSION,
            ))
    return observations
