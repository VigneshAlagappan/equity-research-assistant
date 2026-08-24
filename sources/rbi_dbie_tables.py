"""Parses RBI DBIE (Database on Indian Economy) single-table exports — the
raw downloads under data/raw/_macro/rbi/MoneyAndBanks/ and
.../PricesAndProduction/ (filenames like "13T_<hash>.XLSX", "19AT_<hash>.XLSX").

Each RBI table number has its own bespoke layout: a title row ("No. 13:
..."), an optional unit row, 1-3 header rows describing each column's
period (frequently spanning several merged cells), an index row (sequential
1, 2, 3... or "%" for a growth-rate column), then data rows with a row
label (e.g. "1. SLR Securities") followed by one numeric value per column.
Rows are series; columns are periods — the transpose of
sources/rbi_indicators.py's shape (rows are periods, columns are series).

RBI's own export has a known quirk: short date-like header text (e.g.
"1-May") sometimes gets auto-converted to a full but WRONG-YEAR datetime
(seen as 1930/1931/2015 in these files) rather than staying as text. This
parser never guesses past that corruption — a column is only ingested when
its header can be confidently resolved to a real calendar date from
uncorrupted text; anything else (including "Financial year so far"/"Y-o-Y"/
"X over Y" growth-rate columns, which aren't a single point-in-time period
at all) is skipped and logged, not force-fit into a period.

Rows are always ingested as period_type="dated" (an irregular as-on-this-
date snapshot, not a fixed recurring cadence like weekly/monthly).
"""

from __future__ import annotations

import calendar
import logging
import re
from datetime import datetime
from pathlib import Path

import openpyxl

from sources.macro import MacroNormalizedObservation

logger = logging.getLogger(__name__)

PARSER_VERSION = "rbi-dbie-table-v1-xlsx"

_TITLE_RE = re.compile(r"^No\.\s*(\d+\s*\(?[a-z]?\)?)\s*:\s*(.+)$", re.IGNORECASE)
# Searched within a column's combined header text, not full-matched against
# one cell — real headers mix descriptive words and the date in one string
# (e.g. "Mar. 31, 2026", "As on March 31, 2026").
_MONTH_DAY_SEARCH_RE = re.compile(
    r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s*(\d{1,2})?\b", re.IGNORECASE
)
_YEAR_SEARCH_RE = re.compile(r"\b(20\d{2})\b")
_SKIP_LABEL_RE = re.compile(r"\bover\b|\bso far\b|\by-o-y\b|\bgrowth\b", re.IGNORECASE)
# Deliberately tight, not "any four-digit year since 2000": RBI's export
# corruption (see module docstring) has shown up as datetime cells with
# years like 1930/1931/2015 where the source text almost certainly meant a
# bare "1-May"-style day/month with no year at all — a narrow window around
# this data's actual reporting period keeps those out without also
# rejecting genuine recent dates.
_PLAUSIBLE_YEAR_RANGE = range(2023, 2028)

_MONTH_NAMES = {m.lower(): i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1
)}


def _slugify(label: str) -> str:
    label = label.replace("–", "-").replace("’", "")
    label = re.sub(r"[^A-Za-z0-9]+", "_", label)
    return re.sub(r"_+", "_", label).strip("_").lower()


def _merged_forward_fill(ws) -> dict[tuple[int, int], object]:
    """(row, col) -> value for every cell inside a merged range, taken from
    the range's top-left cell — openpyxl only stores the value there, every
    other cell in the range reads as None otherwise."""
    fill: dict[tuple[int, int], object] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = ws.cell(merged_range.min_row, merged_range.min_col).value
        if top_left is None:
            continue
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                fill[(row, col)] = top_left
    return fill


def _cell_value(ws, merged: dict[tuple[int, int], object], row: int, col: int) -> object:
    value = ws.cell(row, col).value
    return value if value is not None else merged.get((row, col))


def _find_title(ws) -> tuple[str, str, int] | None:
    """(table_tag, title_text, row) from a "No. NN: <title>" cell, e.g.
    ("13", "Scheduled Commercial Banks' Investments", 2). The row is used
    by callers to skip the title (and often-merged-across-columns unit row
    right after it) when building per-column header text — title/unit
    cells frequently span the whole table width, and would otherwise leak
    into every column's reconstructed label."""
    for row in ws.iter_rows(min_row=1, max_row=4):
        for cell in row:
            if isinstance(cell.value, str):
                match = _TITLE_RE.match(cell.value.strip())
                if match:
                    tag = re.sub(r"[^a-z0-9]", "", match.group(1).lower())
                    return tag, match.group(2).strip(), cell.row
    return None


def _find_index_row(ws) -> tuple[int, list[tuple[int, str]]] | None:
    """(row, [(col, unit_marker)]) — the row of sequential 1,2,3.../"%"
    markers, and which columns count as real data columns. unit_marker is
    "%" if that column's index cell literally reads "%", else "".
    """
    for row in ws.iter_rows(min_row=1, max_row=10):
        markers = [(c.column, c.value) for c in row if c.value is not None]
        numeric_markers = [m for _, m in markers if isinstance(m, (int, float))]
        if len(numeric_markers) >= 2 and list(numeric_markers[: len(numeric_markers)]) == list(
            range(1, len(numeric_markers) + 1)
        ):
            columns = [(col, "%" if val == "%" else "") for col, val in markers]
            return row[0].row, columns
    return None


def _year_candidates(header_cells: list[object], combined_text: str) -> list[int]:
    from_text = [int(y) for y in _YEAR_SEARCH_RE.findall(combined_text)]
    from_ints = [int(c) for c in header_cells if isinstance(c, int) and not isinstance(c, bool) and c in _PLAUSIBLE_YEAR_RANGE]
    return [y for y in from_text + from_ints if y in _PLAUSIBLE_YEAR_RANGE]


def _resolve_column_date(header_cells: list[object]) -> str | None:
    """Best-effort date from one column's stacked header cells (title-row
    down to the index row, merged-cell-filled). Returns "YYYY-MM-DD", or
    None if the column isn't confidently a single point-in-time date —
    including a bare year with no month/day anywhere in its header stack,
    which the caller resolves separately via cross-column backfill."""
    strings = [str(t).strip() for t in header_cells if isinstance(t, str) and t.strip()]
    if any(_SKIP_LABEL_RE.search(s) for s in strings):
        return None  # growth/comparison column, not a point-in-time period

    # A datetime cell is only trusted within a tight, plausible year window —
    # RBI's export corruption (module docstring) produces exactly this shape
    # (a real-looking datetime, wildly wrong year) often enough that an
    # untrusted one must fall through to text-based reconstruction instead.
    plausible_dates = [t for t in header_cells if isinstance(t, datetime) and t.year in _PLAUSIBLE_YEAR_RANGE]
    if plausible_dates:
        return plausible_dates[0].strftime("%Y-%m-%d")

    combined_text = " ".join(strings)
    month_match = _MONTH_DAY_SEARCH_RE.search(combined_text)
    if month_match is None:
        return None
    years = _year_candidates(header_cells, combined_text)
    if not years:
        return None
    month = _MONTH_NAMES[month_match.group(1)[:3].lower()]
    year = max(years)  # most recent year mentioned in this column's header stack
    # No explicit day in the header text (e.g. "At End-March", a bare "May")
    # -> anchor to that month's actual last day, not a fixed day-of-month —
    # both this table's real meanings for a bare month ("as of fiscal
    # month-end" / "monthly average, anchored to period-end") are reasonably
    # represented by the calendar-correct month-end, unlike an arbitrary
    # fixed day that's wrong for every month shorter than it.
    day = int(month_match.group(2)) if month_match.group(2) else calendar.monthrange(year, month)[1]
    try:
        return datetime(year, month, day).strftime("%Y-%m-%d")
    except ValueError:
        return None


def _bare_year(header_cells: list[object]) -> int | None:
    """A column whose header stack is just a year, no month/day at all —
    resolved by the caller via cross-column backfill (same as
    _resolve_column_date, this never trusts an implausible-year datetime)."""
    strings = [str(t).strip() for t in header_cells if isinstance(t, str) and t.strip()]
    if any(_SKIP_LABEL_RE.search(s) for s in strings):
        return None
    combined_text = " ".join(strings)
    years = _year_candidates(header_cells, combined_text)
    return max(years) if years else None


def parse_rbi_dbie_table(file_path: Path) -> list[MacroNormalizedObservation]:
    """Parse one RBI DBIE single-table export. Raises ValueError if the
    file doesn't have the expected title-row/index-row shape at all — a
    defensive check, not a claim that every column within a recognized
    table will actually be ingested (ambiguous columns are skipped, not
    guessed)."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    title = _find_title(ws)
    index = _find_index_row(ws)
    if title is None or index is None:
        raise ValueError(f"{file_path} does not look like an RBI DBIE table export (no title/index row found)")
    table_tag, table_title, title_row = title
    index_row, data_columns = index
    merged = _merged_forward_fill(ws)

    # Resolve each data column's period, backfilling bare-year-only columns
    # with the first confidently-resolved sibling's month/day — RBI's
    # "outstanding as on"-style tables are consistently same-day-different-year.
    resolved: dict[int, str] = {}
    unresolved_bare_years: dict[int, int] = {}
    for col, _unit_marker in data_columns:
        header_texts = [_cell_value(ws, merged, r, col) for r in range(title_row + 1, index_row)]
        date = _resolve_column_date(header_texts)
        if date is not None:
            resolved[col] = date
            continue
        year = _bare_year(header_texts)
        if year is not None:
            unresolved_bare_years[col] = year

    if unresolved_bare_years and resolved:
        anchor_month, anchor_day = map(int, next(iter(resolved.values())).split("-")[1:])
        for col, year in unresolved_bare_years.items():
            try:
                resolved[col] = datetime(year, anchor_month, anchor_day).strftime("%Y-%m-%d")
            except ValueError:
                continue

    skipped_columns = len(data_columns) - len(resolved)
    if skipped_columns:
        logger.warning(
            "%s [%s]: could not confidently date %d of %d column(s) — skipped",
            file_path, ws.title, skipped_columns, len(data_columns),
        )

    unit_by_col = {col: unit_marker for col, unit_marker in data_columns}
    observations: list[MacroNormalizedObservation] = []
    for row in ws.iter_rows(min_row=index_row + 1, max_row=ws.max_row):
        label_cell = next((c for c in row if isinstance(c.value, str) and c.value.strip()), None)
        if label_cell is None:
            continue
        series_key = f"t{table_tag}_{_slugify(label_cell.value)}"
        for col, period in resolved.items():
            value = ws.cell(row[0].row, col).value
            if not isinstance(value, (int, float)) or isinstance(value, bool):
                continue
            unit = "PERCENT" if unit_by_col.get(col) == "%" else "INR_CRORE" if "₹" in (table_title or "") or "crore" in (table_title or "").lower() else "NUMBER"
            observations.append(MacroNormalizedObservation(
                series_key=series_key, period_type="dated", period=period,
                value=float(value), unit=unit, source="rbi", source_file=str(file_path),
                parser_version=PARSER_VERSION,
            ))
    return observations


# ------------------------------------------------------------------
# A second, less common RBI DBIE orientation: rows are dates, columns are
# series — the same shape sources/rbi_indicators.py's workbook uses, but as
# a single ad-hoc table export (e.g. "No. 27: Daily Call Money Rates")
# rather than the fixed 4-sheet publication. parse_rbi_dbie_table() above
# assumes the opposite (rows are series) and would misread this — detected
# separately via looks_like_row_oriented_dbie_table() before falling back
# to the column-oriented parser.
# ------------------------------------------------------------------

_DAILY_DATE_RE = re.compile(r"^([A-Za-z]{3,9})\.?\s+(\d{1,2}),?\s*(\d{4})$")


def _parse_daily_date(text: object) -> str | None:
    if not isinstance(text, str):
        return None
    match = _DAILY_DATE_RE.match(text.strip())
    if not match:
        return None
    month_name, day, year = match.groups()
    month = _MONTH_NAMES.get(month_name[:3].lower())
    if month is None:
        return None
    try:
        return datetime(int(year), month, int(day)).strftime("%Y-%m-%d")
    except ValueError:
        return None


def looks_like_row_oriented_dbie_table(file_path: Path) -> bool:
    """True if this table's data rows are dates (column B) rather than row
    labels — checked by trying to parse the first data row's label cell as
    a date, using the same title/index-row anchors parse_rbi_dbie_table()
    itself relies on."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        title = _find_title(ws)
        index = _find_index_row(ws)
        if title is None or index is None:
            return False
        index_row, _data_columns = index
        first_data_row = next(ws.iter_rows(min_row=index_row + 1, max_row=index_row + 1), None)
        if first_data_row is None:
            return False
        label_value = next((c.value for c in first_data_row if isinstance(c.value, str) and c.value.strip()), None)
        return _parse_daily_date(label_value) is not None
    except Exception:
        return False


def parse_rbi_daily_rate_table(file_path: Path) -> list[MacroNormalizedObservation]:
    """Parse a row-oriented single-table RBI DBIE export (dates as rows,
    series as columns) — e.g. "No. 27: Daily Call Money Rates". A column is
    skipped if any of its values are non-numeric (e.g. a "4.50-5.35" rate
    range, not a single number) rather than guessing which end of the range
    to keep."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    title = _find_title(ws)
    index = _find_index_row(ws)
    if title is None or index is None:
        raise ValueError(f"{file_path} does not look like an RBI DBIE table export (no title/index row found)")
    table_tag, table_title, title_row = title
    index_row, data_columns = index
    merged = _merged_forward_fill(ws)
    unit_row_text = str(_cell_value(ws, merged, title_row + 1, 2) or "").lower()
    unit = "PERCENT" if "per cent" in unit_row_text or "%" in unit_row_text else "NUMBER"

    labels: dict[int, str] = {}
    for col, _unit_marker in data_columns:
        # title_row + 2, not + 1: skips both the title row and the
        # unit-description row right after it (e.g. "(Per cent per
        # annum)") — both are frequently merged across the whole table
        # width and would otherwise leak into every column's label.
        header_texts = [t for r in range(title_row + 2, index_row) if (t := _cell_value(ws, merged, r, col)) is not None]
        if header_texts:
            labels[col] = _slugify(" ".join(str(t) for t in header_texts))

    date_rows = [
        (period, row)
        for row in ws.iter_rows(min_row=index_row + 1, max_row=ws.max_row)
        if (period := _parse_daily_date(row[1].value)) is not None
    ]

    # A column only qualifies if every value it has across all date rows is
    # numeric — one non-numeric cell (a "4.50-5.35" rate range) marks the
    # whole column unresolvable, not just that one row.
    qualifying_cols = []
    for col in labels:
        cells = [row[col - 1].value for _period, row in date_rows]
        non_blank = [c for c in cells if c is not None and str(c).strip() != ""]
        if non_blank and all(isinstance(c, (int, float)) and not isinstance(c, bool) for c in non_blank):
            qualifying_cols.append(col)

    skipped = len(labels) - len(qualifying_cols)
    if skipped:
        logger.warning(
            "%s [%s]: skipped %d column(s) with non-numeric values (e.g. a rate range)",
            file_path, ws.title, skipped,
        )

    observations: list[MacroNormalizedObservation] = []
    for period, row in date_rows:
        for col in qualifying_cols:
            value = row[col - 1].value
            if not isinstance(value, (int, float)) or isinstance(value, bool):
                continue
            observations.append(MacroNormalizedObservation(
                series_key=f"t{table_tag}_{labels[col]}", period_type="dated", period=period,
                value=float(value), unit=unit, source="rbi", source_file=str(file_path),
                parser_version=PARSER_VERSION,
            ))
    return observations
