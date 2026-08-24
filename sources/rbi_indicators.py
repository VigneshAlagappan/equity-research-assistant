"""Parses RBI's "50 Macroeconomic Indicators" workbook — a Handbook-of-
Statistics-style publication with four sheets (Weekly/Fortnightly/Monthly/
Quarterly), each a period x indicator matrix covering dozens of series at
once.

Not folded into sources/macro.py's MacroDataAdapter: that adapter's contract
is one CSV file = one series (period,value,unit columns); this is one XLSX
file with many series across four differently-shaped sheets. This module's
layout assumptions are specific to this one RBI publication, not a
generalizable macro-source convention the way the CSV shape is — see
sources/rbi_dbie_tables.py for the (also bespoke, differently-shaped) single-
table RBI DBIE exports.

Period convention: annual/monthly already fit sources/macro.py's existing
period_type vocabulary. Weekly/Fortnightly/Quarterly periods are stored as
the exact reported date ("YYYY-MM-DD", infer_period_type()'s "dated" shape)
rather than inventing a fiscal-vs-calendar "Q1"/"Q2" label the sheet's own
header doesn't state unambiguously.
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path

import openpyxl

from sources.macro import MacroNormalizedObservation

logger = logging.getLogger(__name__)

PARSER_VERSION = "rbi-indicators-v1-xlsx"

# Sheet name -> macro_observations.period_type this sheet's Period column
# should be stored as.
SHEET_PERIOD_TYPES = {
    "Weekly": "weekly",
    "Fortnightly": "fortnightly",
    "Monthly": "monthly",
    "Quarterly": "quarterly",
}

_UNIT_HINTS = [
    (re.compile(r"\(%\)"), "PERCENT"),
    (re.compile(r"US\s*\$\s*Million", re.IGNORECASE), "USD_MILLION"),
    (re.compile(r"₹\s*Crore", re.IGNORECASE), "INR_CRORE"),
]
_MONTH_ABBR_RE = re.compile(r"^([A-Za-z]{3})-(\d{4})$")
# Only a paren whose content looks like a unit gets dropped when building a
# series_key — e.g. "(US $ Million)" — everything else, like "(2012=100)"
# or "(2024=100)", stays: two CPI columns with different index bases must
# not collapse to the same series_key just because both parens get stripped.
_UNIT_PAREN_HINT_RE = re.compile(r"%|million|crore|lakh", re.IGNORECASE)


def _strip_unit_parens(label: str) -> str:
    return re.sub(r"\([^()]*\)", lambda m: "" if _UNIT_PAREN_HINT_RE.search(m.group(0)) else m.group(0), label)


def _slugify(label: str) -> str:
    """"Foreign Exchange Reserves \\n(US $ Million)" -> "foreign_exchange_reserves".
    "Consumer Price Index  (2012=100)" -> "consumer_price_index_2012_100" (kept
    distinct from the "(2024=100)" column, not merged into one series)."""
    label = _strip_unit_parens(label)
    label = label.replace("–", "-").replace("’", "").replace("à", "a")
    label = re.sub(r"[^A-Za-z0-9]+", "_", label)
    return re.sub(r"_+", "_", label).strip("_").lower()


def _unit_for_label(label: str) -> str:
    for pattern, unit in _UNIT_HINTS:
        if pattern.search(label):
            return unit
    return "NUMBER"  # dimensionless index/rate — same default financial_observations already uses


def _period_string(period_type: str, raw_period: object) -> str | None:
    if period_type == "monthly":
        if isinstance(raw_period, str):
            match = _MONTH_ABBR_RE.match(raw_period.strip())
            if match:
                month_abbr, year = match.groups()
                month = datetime.strptime(month_abbr, "%b").month
                return f"{year}-{month:02d}"
        return None
    # weekly / fortnightly / quarterly: the Period column is a real reported date
    if isinstance(raw_period, datetime):
        return raw_period.strftime("%Y-%m-%d")
    return None


def _find_header_row(ws) -> tuple[int, int]:
    """(header_row, first_data_col) — first_data_col is the column right
    after the cell reading exactly "Period"."""
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == "Period":
                return cell.row, cell.column + 1
    raise ValueError(f"{ws.title}: could not find a 'Period' header cell")


def _parse_sheet(ws, period_type: str, file_path: Path) -> list[MacroNormalizedObservation]:
    header_row, first_col = _find_header_row(ws)
    columns = [
        (cell.column, _slugify(str(cell.value)), _unit_for_label(str(cell.value)))
        for cell in ws[header_row][first_col - 1:]
        if cell.value is not None
    ]

    observations: list[MacroNormalizedObservation] = []
    skipped_periods = 0
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        raw_period = row[first_col - 2].value  # one column left of the first data column
        period = _period_string(period_type, raw_period)
        if period is None:
            skipped_periods += 1
            continue
        for col_index, series_key, unit in columns:
            value = row[col_index - 1].value
            if not isinstance(value, (int, float)) or isinstance(value, bool):
                continue  # "-" placeholders and blanks land here too
            observations.append(MacroNormalizedObservation(
                series_key=series_key, period_type=period_type, period=period,
                value=float(value), unit=unit, source="rbi", source_file=str(file_path),
                parser_version=PARSER_VERSION,
            ))

    if skipped_periods:
        logger.warning(
            "%s [%s]: skipped %d row(s) with an unparseable Period value", file_path, ws.title, skipped_periods
        )
    return observations


def looks_like_rbi_indicator_workbook(file_path: Path) -> bool:
    """Cheap, defensive shape check — used to route between this module and
    sources/rbi_dbie_tables.py's single-table parser without guessing."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
    except Exception:
        return False
    return set(SHEET_PERIOD_TYPES).issubset(wb.sheetnames)


def parse_rbi_indicator_workbook(file_path: Path) -> list[MacroNormalizedObservation]:
    """Parse every sheet of the "50 Macroeconomic Indicators" workbook.
    Raises ValueError if the workbook doesn't have the expected sheet names
    — a defensive check against silently mis-parsing a differently-shaped
    file that happens to share the .xlsx extension."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if not set(SHEET_PERIOD_TYPES).issubset(wb.sheetnames):
        raise ValueError(
            f"{file_path} does not look like the RBI '50 Macroeconomic Indicators' workbook "
            f"(expected sheets {sorted(SHEET_PERIOD_TYPES)}, got {wb.sheetnames})"
        )

    observations: list[MacroNormalizedObservation] = []
    for sheet_name, period_type in SHEET_PERIOD_TYPES.items():
        observations.extend(_parse_sheet(wb[sheet_name], period_type, file_path))
    return observations
