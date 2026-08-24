"""ScreenerAdapter — parses Screener.in .xlsx exports via the "Data Sheet" tab.

Verified against a real export (Jio Financial Services): the pretty
per-topic sheets (Profit & Loss, Balance Sheet, Quarters, Cash Flow) are
formula views over "Data Sheet" and read back as empty cells through
openpyxl — their formulas aren't cached as values in a plain export. "Data
Sheet" is Screener's own stable source tab (its own header says "PLEASE DO
NOT MAKE ANY CHANGES TO THIS SHEET"), so that's what this adapter reads.

"Data Sheet" is section-based, not one sheet per topic: a label-only section
header row ("PROFIT & LOSS", "Quarters", "BALANCE SHEET", "CASH FLOW:",
"RATIOS"), immediately followed by a "Report Date" row giving each column's
period-end date as a real date/datetime value (not "Mar-24" text), followed
by metric rows aligned to those same columns until a blank row or the next
section header.

Row-label -> metric_key mapping is entirely data-driven (metric_aliases
table, seeded in normalization/financials.py) — this adapter never hardcodes
a row position or a metric name.
"""

from __future__ import annotations

import datetime as dt
import logging
import sqlite3
from pathlib import Path

import openpyxl

from normalization.financials import build_observations_from_periods
from normalization.periods import fiscal_year_and_quarter_from_date
from sources.base import NormalizedObservation, SourceAdapter

logger = logging.getLogger(__name__)

PARSER_VERSION = "screener-v2-datasheet"

DATA_SHEET_NAME = "Data Sheet"

# Section header text (normalized: stripped, upper-cased, trailing ":" removed)
# -> period_type. RATIOS is included defensively (a bank export may carry a
# distinct Ratios section reporting GNPA/NNPA/CASA/NIM) but hasn't been
# verified against a real bank Data Sheet — only against the NBFC export
# this adapter was built and tested against, which has no Ratios section.
SECTION_PERIOD_TYPES: dict[str, str] = {
    "PROFIT & LOSS": "annual",
    "QUARTERS": "quarterly",
    "BALANCE SHEET": "annual",
    "CASH FLOW": "annual",
    "RATIOS": "annual",
    # No Report Date row of its own — see _parse_rows' fallback, which
    # reuses the prior annual section's period columns for it. Verified
    # against a real ICICI Bank export: carries "Adjusted Equity Shares in
    # Cr", already split/bonus-adjusted and crore-scaled (unlike "No. of
    # Equity Shares" in BALANCE SHEET, a raw share count in a different
    # order of magnitude — deliberately not aliased to the same metric_key,
    # that would reconcile two wildly different scales under one number).
    "DERIVED": "annual",
}

_REPORT_DATE_LABEL = "REPORT DATE"


def _normalize_label(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    return value.strip().upper().rstrip(":")


class ScreenerAdapter(SourceAdapter):
    source_id = "screener"

    def __init__(self, conn: sqlite3.Connection):
        self._conn = conn

    def parse(
        self,
        file_path: Path,
        company_id: str,
        statement_type: str = "consolidated",
        **kwargs: object,
    ) -> list[NormalizedObservation]:
        """Parse every recognized section in the "Data Sheet" tab.

        statement_type applies to the whole export: Screener decides
        consolidated-vs-standalone at export time, not per section, so the
        caller (CLI/pipeline) states which one this file is.
        """
        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        try:
            if DATA_SHEET_NAME not in workbook.sheetnames:
                raise ValueError(
                    f"{file_path} has no {DATA_SHEET_NAME!r} tab — is this a Screener.in export?"
                )
            rows = list(workbook[DATA_SHEET_NAME].iter_rows(values_only=True))
        finally:
            workbook.close()

        return self._parse_rows(rows, company_id=company_id, source_file=str(file_path), statement_type=statement_type)

    def _parse_rows(
        self, rows: list[tuple[object, ...]], *, company_id: str, source_file: str, statement_type: str
    ) -> list[NormalizedObservation]:
        observations: list[NormalizedObservation] = []
        i = 0
        n = len(rows)
        # Some exports carry a trailing section with no "Report Date" row of
        # its own (verified against a real ICICI Bank export: "DERIVED:" is
        # a bare label immediately followed by its data row, "Adjusted
        # Equity Shares in Cr" — no Report Date in between). Its values line
        # up with the last well-formed section's own year columns, so that
        # mapping is reused rather than the section being skipped outright —
        # only when a same-period_type section has already supplied one;
        # a file where the very first section lacks a Report Date row still
        # hits the original warn-and-skip path below, unchanged.
        last_period_by_column: dict[int, str] | None = None
        last_period_type: str | None = None

        while i < n:
            section = _normalize_label(rows[i][0] if rows[i] else None)
            if section not in SECTION_PERIOD_TYPES:
                i += 1
                continue

            period_type = SECTION_PERIOD_TYPES[section]
            section_label = rows[i][0]
            i += 1

            has_report_date = i < n and _normalize_label(rows[i][0] if rows[i] else None) == _REPORT_DATE_LABEL
            if has_report_date:
                period_by_column = self._period_headers_by_column(rows[i], period_type)
                i += 1
            elif last_period_by_column is not None and last_period_type == period_type:
                logger.info(
                    "Section %r has no 'Report Date' row of its own — reusing the last %s "
                    "section's period columns",
                    section_label, period_type,
                )
                period_by_column = last_period_by_column
            else:
                logger.warning(
                    "Section %r has no 'Report Date' row immediately after it — skipping", section_label
                )
                continue

            last_period_by_column, last_period_type = period_by_column, period_type

            while i < n:
                row = rows[i]
                row_label = row[0] if row else None
                if row_label is None:
                    i += 1
                    break  # blank row ends the section
                if _normalize_label(row_label) in SECTION_PERIOD_TYPES:
                    break  # next section header — outer loop handles it, don't consume it here

                period_values = {
                    period: row[col_index]
                    for col_index, period in period_by_column.items()
                    if col_index < len(row)
                }
                observations.extend(
                    build_observations_from_periods(
                        self._conn,
                        company_id=company_id,
                        source=self.source_id,
                        source_file=source_file,
                        parser_version=PARSER_VERSION,
                        period_type=period_type,
                        statement_type=statement_type,
                        row_label=str(row_label).strip(),
                        period_values=period_values,
                    )
                )
                i += 1

        return observations

    @staticmethod
    def _period_headers_by_column(
        date_row: tuple[object, ...], period_type: str
    ) -> dict[int, tuple[str, str | None]]:
        """Map each column index in the "Report Date" row to its (fiscal_year, quarter)."""
        headers: dict[int, tuple[str, str | None]] = {}
        for col_index, cell in enumerate(date_row):
            if not isinstance(cell, dt.datetime):
                continue
            headers[col_index] = fiscal_year_and_quarter_from_date(cell.date(), period_type)
        return headers
